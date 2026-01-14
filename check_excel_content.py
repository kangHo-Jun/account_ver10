import openpyxl
import sys

try:
    wb = openpyxl.load_workbook('1CIY6HQXG9KNMPF.xlsx', data_only=True)
    ws = wb.active
    print(f"Total Rows: {ws.max_row}")
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=40)):
        row_data = [str(cell.value) if cell.value is not None else "" for cell in row]
        print(f"Row {i+1}: {' | '.join(row_data)}")
except Exception as e:
    print(f"Error: {e}")
